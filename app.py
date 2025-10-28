# app.py - CLOUDTECH SLA ANALYZER (Streamlit Cloud Ready)
import streamlit as st
import pandas as pd
import io
import base64
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
import tempfile
import urllib.request
from PIL import Image as PILImage

# === PAGE CONFIG ===
st.set_page_config(page_title="CloudTech SLA Analyzer", layout="centered")

# === DARK MODE + PROFESSIONAL STYLE ===
st.markdown("""
<style>
    .main { background: #0a0e17; color: #fafafa; padding: 2rem; }
    .title { font-size: 2.8rem; font-weight: bold; color: #1E90FF; text-align: center; margin: 1rem 0; }
    .subtitle { text-align: center; color: #aaa; margin-bottom: 2rem; }
    .upload { border: 3px dashed #1E90FF; padding: 40px; border-radius: 12px; text-align: center; margin: 20px 0; }
    .upload:hover { background: #1a1f2e; }
    .metric { background: #1E90FF; color: white; padding: 15px; border-radius: 8px; text-align: center; font-size: 1.2em; margin: 10px 0; }
    .btn { background: #27ae60; color: white; padding: 12px 24px; border: none; border-radius: 6px; cursor: pointer; text-decoration: none; display: inline-block; margin: 10px 0; }
    .footer { text-align: center; color: #888; margin-top: 50px; font-size: 0.9em; }
    .stAlert { background: #2c3e50; border: 1px solid #1E90FF; }
</style>
""", unsafe_allow_html=True)

# === HEADER ===
st.markdown("<div class='title'>CLOUDTECH SLA ANALYZER</div>", unsafe_allow_html=True)
st.markdown("<p class='subtitle'>Upload any Excel → Get instant SLA compliance report</p>", unsafe_allow_html=True)

# === CLOUDTECH LOGO ===
logo_url = "https://i.imgur.com/5v5V5v5.png"  # Replace with your real logo
try:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as f:
        urllib.request.urlretrieve(logo_url, f.name)
        logo_img = PILImage.open(f.name)
        st.image(logo_img, width=180, use_column_width=False)
except Exception as e:
    st.markdown("<h2 style='text-align:center; color:#1E90FF;'>CLOUDTECH</h2>", unsafe_allow_html=True)

# === FILE UPLOAD ===
uploaded_file = st.file_uploader("Drop your Excel file here", type=["xlsx"], key="uploader")

if uploaded_file:
    with st.spinner("Analyzing your data..."):
        # === READ EXCEL & AUTO-DETECT SHEET ===
        try:
            xl = pd.ExcelFile(uploaded_file)
        except Exception as e:
            st.error(f"Invalid Excel file: {e}")
            st.stop()

        df = None
        valid_sheet = None
        for sheet in xl.sheet_names:
            try:
                temp_df = pd.read_excel(uploaded_file, sheet_name=sheet)
                cols = temp_df.columns.str.lower()
                if (cols.str.contains('number').any() and 
                    cols.str.contains('created').any() and 
                    cols.str.contains('actual.*end|work.*end|close', regex=True).any()):
                    df = temp_df.copy()
                    valid_sheet = sheet
                    break
            except:
                continue

        if df is None:
            st.error("No sheet found with required columns: **Number**, **Created**, **Actual work end**")
            st.stop()

        st.success(f"Found data in sheet: **{valid_sheet}** → {len(df)} rows")

        # === AUTO-DETECT COLUMNS ===
        try:
            cols_lower = df.columns.str.lower()
            num_col = df.columns[cols_lower.str.contains('number')][0]
            created_col = df.columns[cols_lower.str.contains('created')][0]
            end_col = df.columns[cols_lower.str.contains('actual.*end|work.*end|close', regex=True)][0]
        except IndexError:
            st.error("Could not detect required columns. Check column names.")
            st.stop()

        df = df[[num_col, created_col, end_col]].copy()
        df.columns = ['Number', 'Created', 'End']

        # === DATE CONVERSION ===
        def to_dt(x):
            if pd.isna(x):
                return pd.NaT
            try:
                if isinstance(x, (int, float)) and x > 10000:
                    return pd.to_datetime(x, unit='D', origin='1899-12-30')
                return pd.to_datetime(x)
            except:
                return pd.NaT

        df['Created'] = df['Created'].apply(to_dt)
        df['End'] = df['End'].apply(to_dt)
        df = df.dropna(subset=['Created', 'End']).copy()

        # === CRITICAL: CHECK FOR EMPTY DATA ===
        if len(df) == 0:
            st.error("No valid date rows found. Check 'Created' and 'Actual work end' columns for correct date format.")
            st.stop()

        # === SLA CALCULATION ===
        df['hours'] = (df['End'] - df['Created']).dt.total_seconds() / 3600
        df['STATUS'] = df['hours'] > 24

        total = len(df)
        within = len(df[~df['STATUS']])
        past = len(df[df['STATUS']])
        avg_h = df['hours'].mean()
        compliance = within / total  # ← Safe: total > 0
        past_pct = past / total

        # === DISPLAY METRICS ===
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"<div class='metric'>Within SLA: <strong>{within}</strong></div>", unsafe_allow_html=True)
            st.markdown(f"<div class='metric'>Avg Close: <strong>{avg_h:.2f}h</strong></div>", unsafe_allow_html=True)
        with col2:
            st.markdown(f"<div class='metric'>Past SLA: <strong>{past}</strong></div>", unsafe_allow_html=True)
            st.markdown(f"<div class='metric'>Compliance: <strong>{compliance:.2%}</strong></div>", unsafe_allow_html=True)

        # === PIE CHART ===
        fig, ax = plt.subplots(figsize=(5, 3))
        ax.pie([within, past], labels=['Within SLA', 'Past SLA'], autopct='%1.2f%%', 
               colors=['#2ecc71', '#e74c3c'], startangle=90)
        ax.axis('equal')
        ax.set_title("SLA Compliance", fontsize=14, color='white')
        plt.tight_layout()
        
        buf = io.BytesIO()
        plt.savefig(buf, format='png', facecolor='#0a0e17', edgecolor='none', bbox_inches='tight')
        buf.seek(0)
        chart_b64 = base64.b64encode(buf.read()).decode()
        st.image(f"data:image/png;base64,{chart_b64}", use_column_width=True)

        # === GENERATE EXCEL REPORT ===
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df[['Number', 'Created', 'End', 'hours', 'STATUS']].to_excel(writer, sheet_name='SLA_Data', index=False)

            # Summary
            summary_data = [
                ['', '', 'Ticket Within SLA', within, ''],
                ['', '', 'Closed past SLA', past, ''],
                ['', '', 'Avg Close (h)', round(avg_h, 2), ''],
                ['', '', 'Past SLA %', past_pct, ''],
                ['', '', 'SLA COMPLIANCE', compliance, '']
            ]
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='SLA_Data', index=False, header=False, startrow=len(df) + 3)

            # === FORMAT EXCEL ===
            wb = writer.book
            ws = writer.sheets['SLA_Data']
            bold = Font(bold=True)
            
            # Bold headers
            for cell in ws[1]:
                cell.font = bold
            
            # Bold summary labels
            for row in range(len(df) + 4, len(df) + 9):
                ws.cell(row=row, column=3).font = bold

            # Format percentages
            ws[f'D{len(df) + 7}'].number_format = '0.00%'  # Past SLA %
            ws[f'D{len(df) + 8}'].number_format = '0.00%'  # Compliance

        output.seek(0)
        excel_b64 = base64.b64encode(output.read()).decode()

        # === DOWNLOAD BUTTON ===
        st.markdown(f'''
        <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_b64}" 
           download="CloudTech_SLA_Report.xlsx" class="btn">
           Download Full Excel Report
        </a>
        ''', unsafe_allow_html=True)

else:
    st.info("Upload your Excel file to generate the SLA report")

# === FOOTER ===
st.markdown("<div class='footer'>© 2025 CloudTech | SLA Analyzer v2.1 | Enterprise Ready</div>", unsafe_allow_html=True)